"""
Instruction Excel Generator — Phase 5
Builds a complete Excel workbook from an ExcelConfig object.
Supports: multiple sheets, formulas, conditional formatting,
          charts, dropdowns, totals rows, styled themes.
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers as xl_numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timezone
from pathlib import Path

from app.services.instruction_parser import ExcelConfig, ColumnSpec, SheetSpec
from app.services.workspace_service import workspace_service
from app.utils.logger import get_logger

logger = get_logger(__name__)

THEMES = {
    "professional": {
        "header_fill": "1F4E79", "header_font": "FFFFFF",
        "alt_row": "EBF3FB", "border": "BDD7EE",
        "totals_fill": "D6E4F0", "accent": "2E75B6",
    },
    "minimal": {
        "header_fill": "404040", "header_font": "FFFFFF",
        "alt_row": "F5F5F5", "border": "DDDDDD",
        "totals_fill": "E8E8E8", "accent": "606060",
    },
    "colorful": {
        "header_fill": "7B2D8B", "header_font": "FFFFFF",
        "alt_row": "F3E8F8", "border": "DDB8E8",
        "totals_fill": "E8D0F0", "accent": "9B59B6",
    },
}

FORMAT_COLORS = {
    "green": ("00AA00", "E2EFDA"),
    "red":   ("CC0000", "FCE4D6"),
    "amber": ("B8860B", "FFF2CC"),
    "bold":  (None,     None),
}

NUMBER_FORMATS = {
    "number":     "#,##0.00",
    "currency":   '"$"#,##0.00',
    "date":       "YYYY-MM-DD",
    "percentage": "0.00%",
    "text":       "@",
}


async def generate_from_instruction(
    config: ExcelConfig,
    session_id: str,
) -> Path:
    """Build a complete workbook from ExcelConfig, return the saved path."""

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    theme = THEMES.get(config.color_theme, THEMES["professional"])

    for sheet_spec in config.sheets:
        ws = wb.create_sheet(sheet_spec.name)
        _build_sheet(ws, sheet_spec, theme, config)

    # Charts — added to a Charts sheet
    if config.charts:
        charts_ws = wb.create_sheet("Charts")
        _add_charts(wb, charts_ws, config, theme)

    # Summary sheet
    if config.has_summary_sheet:
        _add_summary(wb, config, theme)

    # Conditional formatting
    for rule in config.conditional_rules:
        if rule.sheet in wb.sheetnames:
            _apply_conditional_rule(wb[rule.sheet], rule, config)

    # Save
    excels_dir = workspace_service.get_excels()
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_title = config.title.replace(" ", "_").replace("/", "-")[:40]
    output_path = excels_dir / f"{safe_title}_{session_id[:8]}_{ts}.xlsx"
    wb.save(output_path)

    logger.info("instruction_excel_generated", path=str(output_path), sheets=len(config.sheets))
    return output_path


def _build_sheet(ws, spec: SheetSpec, theme: dict, config: ExcelConfig):
    """Build one sheet from a SheetSpec."""
    thin = Side(border_style="thin", color=theme["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header row ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    for col_idx, col in enumerate(spec.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.name)
        cell.font = Font(bold=True, color=theme["header_font"], size=11)
        cell.fill = PatternFill("solid", fgColor=theme["header_fill"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col.width

    # ── Sample rows ───────────────────────────────────────────────────────────
    for row_idx in range(2, 2 + spec.sample_rows):
        for col_idx, col in enumerate(spec.columns, 1):
            value = _get_sample_value(col, row_idx - 1)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=theme["alt_row"])
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if col.type in NUMBER_FORMATS:
                cell.number_format = NUMBER_FORMATS[col.type]

    last_data_row = 1 + spec.sample_rows

    # ── Totals row ────────────────────────────────────────────────────────────
    if spec.has_totals_row:
        totals_row = last_data_row + 1
        ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=totals_row, column=1).fill = PatternFill("solid", fgColor=theme["totals_fill"])

        for col_idx, col in enumerate(spec.columns, 1):
            if col.type in ("number", "currency", "percentage"):
                col_letter = get_column_letter(col_idx)
                formula = f"=SUM({col_letter}2:{col_letter}{last_data_row})"
                cell = ws.cell(row=totals_row, column=col_idx, value=formula)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor=theme["totals_fill"])
                cell.border = border
                if col.type in NUMBER_FORMATS:
                    cell.number_format = NUMBER_FORMATS[col.type]

    # ── Freeze + filter ───────────────────────────────────────────────────────
    if spec.freeze_header:
        ws.freeze_panes = "A2"
    if spec.has_filters:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(spec.columns))}1"


def _get_sample_value(col: ColumnSpec, row_num: int):
    """Return a sensible placeholder value for a column type."""
    if col.formula:
        return col.formula
    samples = {
        "number":     row_num * 100,
        "currency":   round(row_num * 250.50, 2),
        "date":       f"2024-01-{row_num:02d}",
        "percentage": round(row_num * 0.1, 2),
        "text":       f"Sample {col.name} {row_num}",
    }
    return samples.get(col.type, f"Item {row_num}")


def _add_charts(wb: Workbook, charts_ws, config: ExcelConfig, theme: dict):
    chart_row = 1
    for chart_spec in config.charts:
        if chart_spec.sheet not in wb.sheetnames:
            continue
        src_ws = wb[chart_spec.sheet]

        # Find column indices
        headers = [src_ws.cell(1, c).value for c in range(1, src_ws.max_column + 1)]
        try:
            cat_col = headers.index(chart_spec.category_column) + 1
            val_col = headers.index(chart_spec.value_column) + 1
        except ValueError:
            cat_col, val_col = 1, 2

        max_row = src_ws.max_row

        ChartClass = {"bar": BarChart, "line": LineChart, "pie": PieChart}.get(
            chart_spec.chart_type, BarChart
        )
        chart = ChartClass()
        chart.title = chart_spec.title
        chart.style = 10
        chart.width = 20
        chart.height = 12

        data = Reference(src_ws, min_col=val_col, min_row=1, max_row=max_row)
        cats = Reference(src_ws, min_col=cat_col, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        charts_ws.add_chart(chart, f"A{chart_row}")
        chart_row += 20


def _add_summary(wb: Workbook, config: ExcelConfig, theme: dict):
    ws = wb.create_sheet("Summary", 0)  # Insert at beginning
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    ws.cell(row=1, column=1, value=config.title).font = Font(
        bold=True, size=14, color=theme["accent"]
    )
    ws.cell(row=2, column=1, value=config.description).font = Font(italic=True, size=10)

    details = [
        ("Sheets", ", ".join(s.name for s in config.sheets)),
        ("Total Columns", sum(len(s.columns) for s in config.sheets)),
        ("Has Charts", "Yes" if config.charts else "No"),
        ("Theme", config.color_theme.capitalize()),
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for i, (key, val) in enumerate(details, 4):
        ws.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws.cell(row=i, column=2, value=str(val))

    ws.freeze_panes = None


def _apply_conditional_rule(ws, rule, config: ExcelConfig):
    try:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if rule.column not in headers:
            return
        col_idx = headers.index(rule.column) + 1
        col_letter = get_column_letter(col_idx)
        last_row = ws.max_row
        cell_range = f"{col_letter}2:{col_letter}{last_row}"

        font_color, fill_color = FORMAT_COLORS.get(rule.format, (None, None))
        fill = PatternFill("solid", fgColor=fill_color) if fill_color else None
        font = Font(bold=True, color=font_color) if font_color else Font(bold=True)

        op_map = {
            "greater_than": "greaterThan",
            "less_than":    "lessThan",
            "equals":       "equal",
        }
        operator = op_map.get(rule.rule_type, "greaterThan")

        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator=operator, formula=[rule.value], fill=fill, font=font),
        )
    except Exception as e:
        logger.warning("conditional_rule_error", error=str(e))
