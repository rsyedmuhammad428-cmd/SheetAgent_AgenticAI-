"""
SheetAgent AI — Phase 7: excel_generator.py
REPLACE: backend/app/agents/excel_generator.py

Phase 7 Root-Cause Fix:
  The old _can_use_real_data() required 50% column-name overlap between
  Gemini-designed columns and extracted real data columns.
  This almost always FAILED because Gemini renames columns (e.g. "Full Name"
  vs "name", "Employee ID" vs "id") → fell through to hardcoded SAMPLE data.

  Phase 7 Strategy:
  1. real_data is ALWAYS written row-by-row as-is (no column-match check).
     The sheet uses the ACTUAL column names from the data, not Gemini's names.
  2. Gemini design is only used for: theme, charts, conditional rules, summary sheet.
  3. If real_data is empty/None → sample data generated as before (no change).
  4. Fallback SAMPLE pool kept but ONLY used when truly no real data exists.
"""
import logging
import re
import json
from pathlib import Path
from datetime import datetime, timezone, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

from app.services.workspace_service import workspace_service

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 7 SAFETY HELPERS
# Gemini sometimes returns wrong types in JSON (list instead of int, etc.)
# These sanitizers prevent crashes like "'list' object cannot be interpreted as integer"
# ─────────────────────────────────────────────────────────────────────────────

def _safe_int(val, default: int) -> int:
    """Convert any Gemini value to int safely."""
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    if isinstance(val, str):
        try:
            return int(val.strip())
        except (ValueError, AttributeError):
            return default
    if isinstance(val, list):
        # Gemini sometimes returns [8] instead of 8
        if val and isinstance(val[0], (int, float, str)):
            return _safe_int(val[0], default)
    return default


def _safe_str(val, default: str = "") -> str:
    """Convert any value to str safely."""
    if val is None:
        return default
    if isinstance(val, list):
        return ", ".join(str(v) for v in val) if val else default
    return str(val)


def _safe_bool(val, default: bool = True) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, int):
        return bool(val)
    if isinstance(val, str):
        return val.lower() not in ("false", "0", "no", "")
    return default


def _sanitize_design(design: dict) -> dict:
    """
    Sanitize the entire design dict returned by Gemini.
    Fixes type mismatches that cause runtime crashes.
    Called once at the top of generate_excel().
    """
    if not isinstance(design, dict):
        return {"title": "Workbook", "sheets": [], "charts": [], "conditional_rules": []}

    # Top-level string fields
    for key in ("title", "description", "schema_type", "color_theme"):
        design[key] = _safe_str(design.get(key, ""), "")

    # has_summary_sheet
    design["has_summary_sheet"] = _safe_bool(design.get("has_summary_sheet", True))

    # Sheets
    clean_sheets = []
    for sheet in (design.get("sheets") or []):
        if not isinstance(sheet, dict):
            continue
        sheet["name"]           = _safe_str(sheet.get("name", "Data"), "Data")[:31]
        sheet["sample_rows"]    = _safe_int(sheet.get("sample_rows", 8), 8)
        sheet["has_totals_row"] = _safe_bool(sheet.get("has_totals_row", True))
        sheet["has_filters"]    = _safe_bool(sheet.get("has_filters", True))
        sheet["freeze_header"]  = _safe_bool(sheet.get("freeze_header", True))

        clean_cols = []
        for col in (sheet.get("columns") or []):
            if not isinstance(col, dict):
                continue
            col["name"]    = _safe_str(col.get("name", "Column"), "Column")
            col["type"]    = _safe_str(col.get("type", "text"), "text")
            col["width"]   = _safe_int(col.get("width", 18), 18)
            col["formula"] = col.get("formula") if isinstance(col.get("formula"), str) else None
            col["required"] = _safe_bool(col.get("required", False), False)

            dv = col.get("dropdown_values")
            if isinstance(dv, list) and dv:
                col["dropdown_values"] = [str(v) for v in dv]
            else:
                col["dropdown_values"] = None

            clean_cols.append(col)
        sheet["columns"] = clean_cols
        clean_sheets.append(sheet)
    design["sheets"] = clean_sheets

    # Charts
    clean_charts = []
    for chart in (design.get("charts") or []):
        if not isinstance(chart, dict):
            continue
        chart["title"]           = _safe_str(chart.get("title", "Chart"))
        chart["chart_type"]      = _safe_str(chart.get("chart_type", "bar"))
        chart["sheet"]           = _safe_str(chart.get("sheet", ""))
        chart["category_column"] = _safe_str(chart.get("category_column", ""))
        chart["value_column"]    = _safe_str(chart.get("value_column", ""))
        clean_charts.append(chart)
    design["charts"] = clean_charts

    # Conditional rules
    clean_rules = []
    for rule in (design.get("conditional_rules") or []):
        if not isinstance(rule, dict):
            continue
        rule["sheet"]     = _safe_str(rule.get("sheet", ""))
        rule["column"]    = _safe_str(rule.get("column", ""))
        rule["rule_type"] = _safe_str(rule.get("rule_type", "equals"))
        rule["value"]     = _safe_str(rule.get("value", ""))
        rule["format"]    = _safe_str(rule.get("format", "amber"))
        clean_rules.append(rule)
    design["conditional_rules"] = clean_rules

    return design


# ── Color themes (unchanged from Phase 6) ────────────────────────────────────
THEMES = {
    "professional": {
        "header_bg": "1F4E79", "header_fg": "FFFFFF",
        "alt_row": "EBF3FB", "border": "B8CCE4",
        "total_bg": "2E75B6", "total_fg": "FFFFFF",
        "title_color": "1F4E79", "sub_color": "555555",
        "accent": "2E75B6",
    },
    "blue": {
        "header_bg": "003087", "header_fg": "FFFFFF",
        "alt_row": "DCE6F1", "border": "9DC3E6",
        "total_bg": "003087", "total_fg": "FFFFFF",
        "title_color": "003087", "sub_color": "444444",
        "accent": "0070C0",
    },
    "green": {
        "header_bg": "1E5631", "header_fg": "FFFFFF",
        "alt_row": "E2EFDA", "border": "A9D18E",
        "total_bg": "1E5631", "total_fg": "FFFFFF",
        "title_color": "1E5631", "sub_color": "444444",
        "accent": "70AD47",
    },
    "orange": {
        "header_bg": "C55A11", "header_fg": "FFFFFF",
        "alt_row": "FCE4D6", "border": "F4B183",
        "total_bg": "C55A11", "total_fg": "FFFFFF",
        "title_color": "C55A11", "sub_color": "555555",
        "accent": "ED7D31",
    },
    "purple": {
        "header_bg": "5B2C8D", "header_fg": "FFFFFF",
        "alt_row": "E9E1F5", "border": "B4A7D6",
        "total_bg": "5B2C8D", "total_fg": "FFFFFF",
        "title_color": "5B2C8D", "sub_color": "555555",
        "accent": "9B59B6",
    },
    "dark": {
        "header_bg": "1A1A2E", "header_fg": "E0E0E0",
        "alt_row": "2D2D44", "border": "404060",
        "total_bg": "16213E", "total_fg": "FFFFFF",
        "title_color": "1A1A2E", "sub_color": "444444",
        "accent": "0F3460",
    },
    "minimal": {
        "header_bg": "363636", "header_fg": "FFFFFF",
        "alt_row": "F5F5F5", "border": "DDDDDD",
        "total_bg": "363636", "total_fg": "FFFFFF",
        "title_color": "363636", "sub_color": "666666",
        "accent": "595959",
    },
}

NUM_FMT = {
    "number":     "#,##0.00",
    "currency":   '"$"#,##0.00',
    "date":       "YYYY-MM-DD",
    "percentage": "0.00%",
    "text":       "@",
    "formula":    "#,##0.00",
}

COND_COLORS = {
    "green": {"font": "1A7A1A", "fill": "C6EFCE"},
    "red":   {"font": "9C0006", "fill": "FFC7CE"},
    "amber": {"font": "9C5700", "fill": "FFEB9C"},
}

# Sample pool — used ONLY when real_data is None/empty
SAMPLE = {
    "names": ["Alice Chen", "Bob Nguyen", "Carlos Diaz", "Diana Patel", "Ethan Brooks",
              "Fatima Malik", "George Kim", "Hannah Johansson", "Ivan Petrov", "Julia Santos",
              "Kevin Osei", "Laura Ferreira", "Mohammed Al-Rashid", "Nina Kowalski", "Oscar Mensah"],
    "departments": ["Engineering", "Sales", "HR", "Finance", "Operations", "Marketing", "IT", "Legal"],
    "products": ["Premium Package", "Basic Plan", "Enterprise Suite", "Starter Kit",
                 "Professional Bundle", "Advanced Module", "Core Service", "Plus Edition"],
    "regions": ["North", "South", "East", "West", "Central", "International"],
    "categories": ["Type A", "Type B", "Type C", "Premium", "Standard", "Economy"],
    "months": ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"],
}


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

async def generate_excel(design: dict, session_id: str, real_data: list | str | None = None) -> Path:
    """
    Generate a professional Excel file.

    Phase 7 change: if real_data is provided, the first data sheet is written
    directly from real_data using the ACTUAL column keys — no column-name
    matching check that can fail.
    """
    # ── Real Data Pre-parsing Engine ─────────────────────────────────────────
    # Safely convert raw text-based JSON/markdown strings into standard structured arrays
    if isinstance(real_data, str):
        real_data_str = real_data.strip()
        try:
            if real_data_str.startswith("[") and real_data_str.endswith("]"):
                real_data = json.loads(real_data_str)
        except Exception:
            pass

    # ── Sanitize Gemini output FIRST — prevents type errors downstream ────────
    design = _sanitize_design(design)

    wb = Workbook()
    wb.remove(wb.active)

    theme_name = design.get("color_theme", "professional")
    theme      = THEMES.get(theme_name, THEMES["professional"])

    has_real_data = bool(real_data and len(real_data) > 0)

    # Summary sheet
    if design.get("has_summary_sheet", True):
        _build_summary_sheet(wb, design, theme)

    sheets = design.get("sheets", [])

    if has_real_data:
        # ── Phase 7 BYPASS MODE: real data → write directly, skip Gemini columns ─
        # This is the ROOT CAUSE fix:
        # Gemini designs column names like "ID, Name, Category, Value" even when
        # the real data has "Name, Score 1, Score 2, Total, Grade".
        # Solution: when real_data exists, use the ACTUAL keys as headers always.
        if isinstance(real_data, list) and len(real_data) > 0 and isinstance(real_data[0], dict):
            real_headers = list(real_data[0].keys())
        else:
            real_headers = []

        if real_headers:
            sheet_name = sheets[0]["name"] if sheets else "Data"
            _build_real_data_sheet(wb, sheet_name, real_headers, real_data, theme, design=design)
            _build_analytics_dashboard(wb, real_data, real_headers, theme)
            logger.info(f"[Phase7] Real data sheet: {len(real_data)} rows × {len(real_headers)} cols — headers: {real_headers}")

            # Remaining sheets use sample data
            for sheet_spec in sheets[1:]:
                _build_data_sheet(wb, sheet_spec, theme, design, real_data=None)
        else:
            sheet_name = sheets[0]["name"] if sheets else "Data"
            if isinstance(real_data, list):
                _build_list_data_sheet(wb, sheet_name, real_data, theme)
            else:
                # If structure is single item string/dict, normalize into a row matrix
                _build_list_data_sheet(wb, sheet_name, [[str(real_data)]], theme)
            
            for sheet_spec in sheets[1:]:
                _build_data_sheet(wb, sheet_spec, theme, design, real_data=None)
    else:
        # ── No real data — generate sample data as Phase 6 did ───────────────
        for sheet_spec in sheets:
            _build_data_sheet(wb, sheet_spec, theme, design, real_data=None)

    # Charts sheet
    if design.get("charts"):
        _build_charts_sheet(wb, design, design["charts"])

    # Conditional formatting
    for rule in design.get("conditional_rules", []):
        sheet_name = rule.get("sheet", "")
        if sheet_name in wb.sheetnames:
            _apply_conditional_format(wb[sheet_name], rule)

    if "Summary" in wb.sheetnames:
        wb.active = wb["Summary"]

    # Save
    excels_dir = workspace_service.get_excels()
    ts         = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    raw_title  = design.get("title", "Workbook")
    safe_title = "".join(c for c in raw_title if c.isalnum() or c in " _-")[:45].strip().replace(" ", "_")
    filename   = f"{safe_title}_{session_id[:8]}_{ts}.xlsx"
    output_path = excels_dir / filename
    wb.save(output_path)

    logger.info(f"[Phase7] Excel saved: {filename}")
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 7 NEW: Write real data directly — no column matching
# ─────────────────────────────────────────────────────────────────────────────

def _build_real_data_sheet(
    wb: Workbook,
    sheet_name: str,
    headers: list,
    real_data: list,
    theme: dict,
    design: dict = None,
):
    """
    Write a sheet using the EXACT keys from real_data rows.
    No column remapping. No sample data. No name matching.
    Every value comes directly from the extracted data.
    Adds heading title row if design["heading_title"] is set.
    """
    ws = wb.create_sheet(sheet_name)
    thin = Side(border_style="thin", color=theme["border"])
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Heading title row (merged, large font) ────────────────────────────────
    heading_title = design.get("heading_title", "") if design else ""
    num_cols      = len(headers)
    header_start_row = 1

    if heading_title and num_cols > 0:
        ws.row_dimensions[1].height = 40
        hc = ws.cell(row=1, column=1, value=heading_title)
        hc.font      = Font(bold=True, size=16, color="FFFFFF", name="Calibri")  # always white on dark bg
        hc.fill      = PatternFill("solid", fgColor=theme["header_bg"])
        hc.alignment = Alignment(horizontal="center", vertical="center")
        if num_cols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        # Separator row
        ws.row_dimensions[2].height = 4
        for ci2 in range(1, num_cols + 1):
            ws.cell(row=2, column=ci2).fill = PatternFill("solid", fgColor=theme["accent"])
        header_start_row = 3
    
    # Store layout metadata so other functions can find headers and data
    ws._sheetagent_header_row  = header_start_row
    ws._sheetagent_data_start  = header_start_row + 1

    # ── Column header row ─────────────────────────────────────────────────────
    ws.row_dimensions[header_start_row].height = 32
    for ci, header in enumerate(headers, 1):
        c = ws.cell(row=header_start_row, column=ci, value=str(header))
        c.font      = Font(bold=True, color=theme["header_fg"], size=11, name="Calibri")
        c.fill      = PatternFill("solid", fgColor=theme["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = brd
        ws.column_dimensions[get_column_letter(ci)].width = max(len(str(header)) + 6, 14)
        
    # ── Data rows ─────────────────────────────────────────────────────────────
    numeric_cols: set[int] = set()
    percentage_cols: set[int] = set()
    data_start_row = header_start_row + 1

    for row_i, row_dict in enumerate(real_data, start=data_start_row):
        ws.row_dimensions[row_i].height = 20
        for ci, header in enumerate(headers, 1):
            val = row_dict.get(header, "") if isinstance(row_dict, dict) else ""
            if val is None:
                val = ""

            # Check string indicators before casting away symbols
            val_str = str(val).strip()
            is_pct = val_str.endswith("%")

            # Auto-detect numeric values
            if isinstance(val, (int, float)):
                numeric_cols.add(ci)
            else:
                # Clean up numeric indicators (, or %) to ensure clean numeric coercion
                clean_str = val_str.replace(",", "").replace("%", "").strip()
                try:
                    if clean_str.lstrip("-").isdigit():
                        val = int(clean_str)
                    else:
                        val = float(clean_str)
                    
                    if is_pct:
                        # Convert whole percentage strings (e.g. "84.3%") to decimal notation (0.843)
                        val = val / 100.0 if val > 1.0 else val
                        percentage_cols.add(ci)
                    else:
                        numeric_cols.add(ci)
                except (ValueError, TypeError):
                    pass  # Keep as basic string object

            cell = ws.cell(row=row_i, column=ci, value=val)
            if row_i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=theme["alt_row"])
            cell.alignment = Alignment(vertical="center")
            cell.border    = brd
            
            # Explicit formatting enforcement on real values
            if ci in percentage_cols:
                cell.number_format = NUM_FMT["percentage"]
            elif ci in numeric_cols:
                if "total" in str(header).lower() or "score" in str(header).lower() or "mark" in str(header).lower():
                    cell.number_format = "#,##0"  # Whole scores don't need trailing cents decimals
                else:
                    cell.number_format = "General"

    last_data_row = data_start_row + len(real_data) - 1

    # ── Totals row ─────────────────────────────────────────────────────────────
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 26
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=total_row, column=ci)
        c.font      = Font(bold=True, color=theme["total_fg"], size=11, name="Calibri")
        c.fill      = PatternFill("solid", fgColor=theme["total_bg"])
        c.border    = brd
        c.alignment = Alignment(horizontal="center", vertical="center")
        
        header_lower = str(headers[ci-1]).lower() if ci <= len(headers) else ""
        
        skip_sum_words = ["year", "date", "month", "id", "code", "serial", "phone", "zip", "age", "rank"]
        should_skip_sum = any(word in header_lower for word in skip_sum_words)
        
        if ci == 1:
            c.value = "TOTAL / SUMMARY"
        elif should_skip_sum:
            pass # Leave blank
        elif ci in percentage_cols or "percentage" in header_lower or "pct" in header_lower:
            col_letter = get_column_letter(ci)
            c.value         = f"=AVERAGE({col_letter}{data_start_row}:{col_letter}{last_data_row})"
            c.number_format = NUM_FMT["percentage"]
        elif ci in numeric_cols:
            col_letter = get_column_letter(ci)
            c.value         = f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})"
            if "total" in header_lower or "score" in header_lower or "mark" in header_lower:
                c.number_format = "#,##0"
            else:
                c.number_format = "#,##0.00"

    # Freeze below column headers; auto-filter on column header row
    freeze_row = header_start_row + 1
    ws.freeze_panes     = f"A{freeze_row}"
    ws.auto_filter.ref  = f"A{header_start_row}:{get_column_letter(len(headers))}{header_start_row}"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:1"


def _build_list_data_sheet(wb: Workbook, sheet_name: str, data: list, theme: dict):
    """Write list-of-lists data (no dict keys)."""
    if not data:
        return
    ws   = wb.create_sheet(sheet_name)
    thin = Side(border_style="thin", color=theme["border"])
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Auto-generate headers Col_1, Col_2...
    num_cols = max(len(row) for row in data if isinstance(row, (list, tuple)))
    headers  = [f"Col_{i}" for i in range(1, num_cols + 1)]

    ws.row_dimensions[1].height = 32
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color=theme["header_fg"], size=11, name="Calibri")
        c.fill = PatternFill("solid", fgColor=theme["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
        ws.column_dimensions[get_column_letter(ci)].width = 15

    for row_i, row in enumerate(data, start=2):
        ws.row_dimensions[row_i].height = 20
        for ci, val in enumerate(row if isinstance(row, (list, tuple)) else [row], start=1):
            cell = ws.cell(row=row_i, column=ci, value=val)
            if row_i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=theme["alt_row"])
            cell.border    = brd
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"


# ─────────────────────────────────────────────────────────────────────────────
# UNCHANGED Phase 6 functions below (sample data path only)
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: Workbook, design: dict, theme: dict):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50

    thin = Side(border_style="thin", color=theme["border"])

    ws.row_dimensions[1].height = 45
    c = ws.cell(row=1, column=1, value=design.get("title", "Workbook"))
    c.font = Font(bold=True, size=18, color=theme["title_color"], name="Calibri")
    c.alignment = Alignment(vertical="center", horizontal="left")
    ws.merge_cells("A1:B1")

    ws.row_dimensions[2].height = 30
    d = ws.cell(row=2, column=1, value=design.get("description", ""))
    d.font = Font(italic=True, size=10, color=theme["sub_color"], name="Calibri")
    d.alignment = Alignment(vertical="center", wrap_text=True)
    ws.merge_cells("A2:B2")

    ws.row_dimensions[3].height = 8
    for col in range(1, 3):
        c = ws.cell(row=3, column=col)
        c.fill = PatternFill("solid", fgColor=theme["accent"])

    data_sheets = design.get("sheets", [])
    details = [
        ("Contents",      ", ".join(s["name"] for s in data_sheets)),
        ("Total Columns", str(sum(len(s.get("columns", [])) for s in data_sheets))),
        ("Charts",        str(len(design.get("charts", []))) or "None"),
        ("Color Theme",   design.get("color_theme", "professional").capitalize()),
        ("Generated",     datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    
    _detail_brd = Border(
        left=Side(border_style="thin", color=theme["border"]),
        right=Side(border_style="thin", color=theme["border"]),
        top=Side(border_style="thin", color=theme["border"]),
        bottom=Side(border_style="thin", color=theme["border"]),
    )
    for i, (key, val) in enumerate(details, 5):
        ws.row_dimensions[i].height = 24
        k = ws.cell(row=i, column=1, value=key)
        k.font      = Font(bold=True, size=10, name="Calibri", color=theme["title_color"])
        k.fill      = PatternFill("solid", fgColor="F0F4F8")
        k.border    = _detail_brd
        k.alignment = Alignment(vertical="center")

        v = ws.cell(row=i, column=2, value=val)
        v.font      = Font(size=10, name="Calibri")
        v.border    = _detail_brd
        v.alignment = Alignment(vertical="center")

    row = len(details) + 6
    ws.row_dimensions[row].height = 20
    nav = ws.cell(row=row, column=1, value="↓ Navigate to sheets below ↓")
    nav.font = Font(italic=True, size=9, color="888888")

    for i, sheet_spec in enumerate(data_sheets, row + 1):
        ws.row_dimensions[i].height = 20
        link = ws.cell(row=i, column=1, value=f"→ {sheet_spec['name']}")
        link.font = Font(size=10, color=theme["accent"], name="Calibri", underline="single")


def _build_data_sheet(wb: Workbook, spec: dict, theme: dict, design: dict, real_data: list | None):
    """Phase 6 data sheet — only used when real_data is None (sample data path)."""
    sheet_name = spec.get("name", "Data")
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = True

    columns = spec.get("columns", [])
    if not columns:
        return

    thin = Side(border_style="thin", color=theme["border"])
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    schema_type = design.get("schema_type", "generic")

    ws.row_dimensions[1].height = 32
    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=col["name"])
        c.font      = Font(bold=True, color=theme["header_fg"], size=11, name="Calibri")
        c.fill      = PatternFill("solid", fgColor=theme["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = brd
        ws.column_dimensions[get_column_letter(ci)].width = max(_safe_int(col.get("width", 18), 18), len(col["name"]) + 4)

    num_rows = _safe_int(spec.get("sample_rows", 8), 8)

    for i in range(num_rows):
        row_idx = i + 2
        row_num = i + 1
        ws.row_dimensions[row_idx].height = 20

        for ci, col in enumerate(columns, 1):
            col_type         = col.get("type", "text")
            formula_template = col.get("formula")

            if formula_template:
                value = _adjust_formula(formula_template, row_idx)
            else:
                value = _generate_sample(col, row_num, schema_type)

            c = ws.cell(row=row_idx, column=ci, value=value)
            if row_idx % 2 == 0:
                c.fill = PatternFill("solid", fgColor=theme["alt_row"])
            c.alignment = Alignment(vertical="center")
            c.border    = brd
            if col_type in NUM_FMT:
                c.number_format = NUM_FMT[col_type]

            dv_values = col.get("dropdown_values")
            if dv_values:
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(str(v) for v in dv_values)}"',
                    showDropDown=False,
                    showErrorMessage=True,
                    errorTitle="Invalid",
                    error="Please select from the list",
                )
                ws.add_data_validation(dv)
                dv.sqref = f"{get_column_letter(ci)}{row_idx}"

    last_data_row = num_rows + 1

    if spec.get("has_totals_row", True):
        tr = last_data_row + 1
        ws.row_dimensions[tr].height = 26
        for ci, col in enumerate(columns, 1):
            c = ws.cell(row=tr, column=ci)
            c.font      = Font(bold=True, color=theme["total_fg"], size=11, name="Calibri")
            c.fill      = PatternFill("solid", fgColor=theme["total_bg"])
            c.border    = brd
            c.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 1:
                c.value = "TOTAL / SUMMARY"
            else:
                col_type   = col.get("type", "text")
                col_letter = get_column_letter(ci)
                if col_type in ("number", "currency") and not col.get("formula"):
                    c.value         = f"=SUM({col_letter}2:{col_letter}{last_data_row})"
                    c.number_format = NUM_FMT.get(col_type, "General")
                elif col_type == "percentage" and not col.get("formula"):
                    c.value         = f"=AVERAGE({col_letter}2:{col_letter}{last_data_row})"
                    c.number_format = NUM_FMT["percentage"]

    if spec.get("freeze_header", True):
        ws.freeze_panes = "A2"
    if spec.get("has_filters", True):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows      = "1:1"


def _adjust_formula(template: str, row: int) -> str:
    return re.sub(r'([A-Z]+)(\d+)', lambda m: f"{m.group(1)}{row}", template)


def _generate_sample(col: dict, idx: int, schema_type: str):
    """Generate realistic sample data — ONLY used when no real_data provided."""
    name     = col["name"].lower()
    col_type = col.get("type", "text")

    if any(w in name for w in [" id", "no.", "number", "ref", "code", "sku"]) or name.endswith("id"):
        prefix = "".join(c for c in col["name"].upper() if c.isalpha())[:3]
        return f"{prefix}-{1000 + idx:04d}"

    if any(w in name for w in ["name", "employee", "student", "client", "customer", "salesperson", "teacher"]):
        return SAMPLE["names"][(idx - 1) % len(SAMPLE["names"])]

    if any(w in name for w in ["department", "dept", "division"]):
        return SAMPLE["departments"][(idx - 1) % len(SAMPLE["departments"])]

    if any(w in name for w in ["title", "position", "role", "designation"]):
        titles = ["Senior Manager", "Analyst", "Team Lead", "Engineer", "Director",
                  "Coordinator", "Specialist", "Executive", "Officer", "Supervisor"]
        return titles[(idx - 1) % len(titles)]

    if any(w in name for w in ["product", "item", "service", "plan", "package"]):
        return SAMPLE["products"][(idx - 1) % len(SAMPLE["products"])]

    if any(w in name for w in ["region", "area", "zone", "territory", "location", "city", "branch"]):
        return SAMPLE["regions"][(idx - 1) % len(SAMPLE["regions"])]

    if col_type == "date" or any(w in name for w in ["date", "dob", "birth", "join", "due", "created", "hired"]):
        d = date(2024, 1, 1) + timedelta(days=(idx - 1) * 12)
        return d.strftime("%Y-%m-%d")

    if "email" in name:
        first = SAMPLE["names"][(idx-1) % len(SAMPLE["names"])].split()[0].lower()
        return f"{first}.{idx}@company.com"

    if any(w in name for w in ["phone", "mobile", "contact", "tel"]):
        return f"+1-{200+idx:03d}-{1000+idx*7:04d}"

    if col_type == "currency":
        ranges = {
            "salary": [3200, 4500, 6800, 5200, 7500, 4200, 8900, 3800, 6200, 5500],
            "basic":  [2800, 3500, 4200, 5800, 3900, 6500, 4800, 3200, 5100, 4400],
            "revenue": [12500, 18300, 9800, 22000, 15400, 11200, 19700, 8900, 16300, 21500],
            "price": [29.99, 49.99, 99.99, 149.99, 19.99, 79.99, 199.99, 39.99, 89.99, 59.99],
            "amount": [1200, 2500, 3800, 1500, 4200, 2100, 3300, 1800, 2700, 3900],
            "target": [15000, 20000, 12000, 25000, 18000, 14000, 22000, 16000, 19000, 21000],
            "cost":   [800, 1200, 2100, 950, 1800, 1050, 2400, 900, 1600, 1350],
        }
        for key, vals in ranges.items():
            if key in name:
                return vals[(idx - 1) % len(vals)]
        return round(1000 + (idx * 573.7 % 9000), 2)

    if col_type == "number":
        if any(w in name for w in ["score", "mark", "grade_num", "points"]):
            return [85, 92, 70, 88, 95, 78, 65, 91, 82, 75][(idx - 1) % 10]
        if any(w in name for w in ["age", "year", "experience"]):
            return 22 + (idx % 20)
        if any(w in name for w in ["qty", "quantity", "units", "stock", "count"]):
            return [10, 25, 5, 42, 18, 7, 33, 60, 15, 28][(idx-1) % 10]
        return idx * 10

    if col_type == "percentage":
        return [0.85, 0.92, 0.70, 0.88, 0.95, 0.78, 1.05, 0.91, 0.82, 0.75][(idx-1) % 10]

    if "status" in name:
        return ["Active", "Active", "Inactive", "Pending", "Active"][(idx-1) % 5]

    if "grade" in name:
        return ["A+", "A", "B+", "B", "C", "A", "B+", "A+", "B", "C"][(idx-1) % 10]

    if "category" in name or "type" in name:
        return SAMPLE["categories"][(idx-1) % len(SAMPLE["categories"])]

    if "month" in name:
        return SAMPLE["months"][(idx-1) % 12]

    if any(w in name for w in ["note", "remark", "comment", "description", "detail"]):
        return ""

    return f"{col['name']} {idx}"


def _build_charts_sheet(wb: Workbook, design: dict, charts: list):
    cws = wb.create_sheet("Charts")
    cws.sheet_view.showGridLines = False
    row = 1

    for cs in charts:
        src_name = cs.get("sheet", "")
        if src_name not in wb.sheetnames:
            data_sheets = [s for s in wb.sheetnames if s not in ("Summary", "Charts", "Dashboard")]
            if not data_sheets:
                continue
            src_name = data_sheets[0]

        src = wb[src_name]

        # Use stored metadata so we read the actual header row, not the heading row
        header_row = getattr(src, "_sheetagent_header_row", 1)
        data_start = getattr(src, "_sheetagent_data_start", 2)

        headers = [src.cell(header_row, c).value
                   for c in range(1, src.max_column + 1)]

        cat_req = str(cs.get("category_column", "")).strip()
        val_req = str(cs.get("value_column", "")).strip()

        # Calculate last real data row (exclude total/summary rows)
        max_data_row = data_start
        for r in range(src.max_row, data_start - 1, -1):
            cell_val = str(src.cell(r, 1).value or "").strip().upper()
            if cell_val and cell_val not in ("TOTAL / SUMMARY", "TOTAL", "SUMMARY", "GRAND TOTAL"):
                max_data_row = r
                break

        if cat_req == "__HEADERS__" and val_req == "__TOTALS__":
            chart_type_raw = str(cs.get("chart_type", "bar")).lower()
            if "pie" in chart_type_raw:
                ChartCls = PieChart
                chart_type = "pie"
            elif "line" in chart_type_raw:
                ChartCls = LineChart
                chart_type = "line"
            else:
                ChartCls = BarChart
                chart_type = "bar"

            chart         = ChartCls()
            chart.title   = cs.get("title", "Chart")
            chart.style   = 10
            chart.width   = 24
            chart.height  = 14

            numeric_cols = []
            for i, h in enumerate(headers, 1):
                if not h: continue
                h_str = str(h).lower()
                if any(x in h_str for x in ["id", "code", "no", "num", "name", "student", "employee", "total"]):
                    continue
                v = src.cell(data_start, i).value
                if isinstance(v, (int, float)):
                    numeric_cols.append(i)

            if not numeric_cols:
                logger.warning("[Chart] No numeric columns found for horizontal chart — skipping")
                continue

            start_col = numeric_cols[0]
            end_col = numeric_cols[-1]

            cats = Reference(src, min_col=start_col, max_col=end_col, min_row=header_row, max_row=header_row)
            
            total_row = max_data_row + 1
            cell_val = str(src.cell(total_row, 1).value or "").strip().upper()
            if cell_val in ("TOTAL / SUMMARY", "TOTAL", "SUMMARY", "GRAND TOTAL"):
                val_row = total_row
            else:
                val_row = max_data_row
                
            data = Reference(src, min_col=start_col, max_col=end_col, min_row=val_row, max_row=val_row)

            chart.add_data(data, titles_from_data=False, from_rows=True)
            chart.set_categories(cats)
            cws.add_chart(chart, f"A{row}")
            row += 24
            logger.info(f"[Chart] Built {chart_type} horizontal chart for headers vs totals")
            continue

        # Find category and value columns by name (case-insensitive)
        cat_col = None
        val_col = None
        for i, h in enumerate(headers, 1):
            if not h:
                continue
            if str(h).strip().lower() == cat_req.lower():
                cat_col = i
            if str(h).strip().lower() == val_req.lower():
                val_col = i

        # Fallbacks: skip ID-like columns (contain "id", "code", "no", "num")
        if cat_col is None:
            for i, h in enumerate(headers, 1):
                if h and not any(x in str(h).lower() for x in ["id", "code", "_no", "num", "ref", "key"]):
                    cat_col = i
                    break
        if cat_col is None:
            cat_col = 1

        if val_col is None:
            # Find first numeric column that isn't the category column
            for r in range(data_start, min(data_start + 5, src.max_row + 1)):
                for i, h in enumerate(headers, 1):
                    if i == cat_col:
                        continue
                    v = src.cell(r, i).value
                    if isinstance(v, (int, float)):
                        val_col = i
                        break
                if val_col:
                    break
        if val_col is None:
            val_col = 2 if cat_col != 2 else 3


        # Count unique category values — skip chart if too many (looks terrible)
        cat_vals = set()
        for r in range(data_start, max_data_row + 1):
            v = src.cell(r, cat_col).value
            if v:
                cat_vals.add(str(v).strip())
        unique_cats = len(cat_vals)

        if unique_cats == 0:
            logger.warning(f"[Chart] No category values found — skipping chart")
            continue
        if unique_cats > 50:
            logger.warning(f"[Chart] {unique_cats} categories too many for chart — skipping")
            continue

        chart_type_raw = str(cs.get("chart_type", "bar")).lower()
        if "pie" in chart_type_raw:
            ChartCls = PieChart
            chart_type = "pie"
        elif "line" in chart_type_raw:
            ChartCls = LineChart
            chart_type = "line"
        else:
            ChartCls = BarChart
            chart_type = "bar"

        chart         = ChartCls()
        chart.title   = cs.get("title", "Chart")
        chart.style   = 10
        chart.width   = 24
        chart.height  = 14

        # Reference value column — header row for title, then data rows
        data = Reference(src, min_col=val_col,
                         min_row=header_row, max_row=max_data_row)
        # Reference category column — data rows only (no header in cats)
        cats = Reference(src, min_col=cat_col,
                         min_row=data_start, max_row=max_data_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        cws.add_chart(chart, f"A{row}")
        row += 24
        logger.info(f"[Chart] Built {chart_type} chart: {cs.get('category_column')} vs {cs.get('value_column')} ({unique_cats} cats)")


def _apply_conditional_format(ws, rule: dict):
    try:
        # Use stored row metadata (set by _build_real_data_sheet)
        # Falls back to row 1 if metadata not present (e.g. old sheets)
        header_row = getattr(ws, "_sheetagent_header_row", 1)
        data_start = getattr(ws, "_sheetagent_data_start", 2)

        # Read actual column headers from the correct header row
        headers = [ws.cell(header_row, c).value
                   for c in range(1, ws.max_column + 1)]

        col_name = rule.get("column", "")
        # Case-insensitive column lookup
        matched_col = None
        for i, h in enumerate(headers, 1):
            if h and str(h).strip().lower() == col_name.strip().lower():
                matched_col = i
                break
        if matched_col is None:
            logger.debug(f"Cond format: column '{col_name}' not found in {headers}")
            return

        cl = get_column_letter(matched_col)

        # Calculate last data row (exclude totals row and empty rows)
        last_data_row = ws.max_row
        # Walk back to find last non-empty row that isn't TOTAL/SUMMARY
        for r in range(ws.max_row, data_start - 1, -1):
            cell_val = str(ws.cell(r, 1).value or "").strip().upper()
            if cell_val and cell_val not in ("TOTAL / SUMMARY", "TOTAL", "SUMMARY", "GRAND TOTAL"):
                last_data_row = r
                break

        if last_data_row < data_start:
            return

        cell_range = f"{cl}{data_start}:{cl}{last_data_row}"

        colors = COND_COLORS.get(rule.get("format", "amber"), COND_COLORS["amber"])
        fill   = PatternFill("solid", fgColor=colors["fill"])
        font   = Font(color=colors["font"], bold=True)

        op_map = {"greater_than": "greaterThan", "less_than": "lessThan", "equals": "equal"}
        op     = op_map.get(rule.get("rule_type", "equals"), "equal")
        val    = rule.get("value", "")

        rule_type = rule.get("rule_type", "equals")
        if rule_type == "equals":
            formula_val = f'"{val}"'
        else:
            try:
                formula_val = str(float(val))
            except (ValueError, TypeError):
                formula_val = str(val)

        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator=op, formula=[formula_val], fill=fill, font=font),
        )
        logger.info(f"[Format] Applied {rule_type} '{val}' → {colors['fill']} on {cell_range}")
    except Exception as e:
        logger.warning(f"Conditional format error: {e}")

def _safe_float(val):
    try:
        return float(str(val).replace(",", "").replace("$", "").replace("%", "").strip())
    except (ValueError, TypeError):
        return 0.0

def _build_analytics_dashboard(wb: Workbook, real_data: list, headers: list, theme: dict):
    if not real_data: return
    headers_lower = [str(h).lower() for h in headers]
    
    finance_cols = [h for h in headers_lower if "revenue" in h or "profit" in h or "sales" in h or "margin" in h or "cost" in h]
    performance_cols = [h for h in headers_lower if "score" in h or "mark" in h or "rating" in h or "grade" in h or "gpa" in h]
    
    is_financial = len(finance_cols) > 0
    is_performance = len(performance_cols) > 0 and not is_financial
    
    if not is_financial and not is_performance:
        _id_like_words  = {"id", "no", "ref", "code", "num", "number", "reference"}
        _id_like_substr = ("serial", "uuid", "zipcode", "phone")
        numeric_cols = []
        for h in headers:
            h_lower = str(h).strip().lower()
            words = re.split(r"[^a-z0-9]+", h_lower)
            if any(w in _id_like_words for w in words if w) or any(s in h_lower for s in _id_like_substr):
                continue
            for row in real_data[:5]:
                val = row.get(h)
                if isinstance(val, (int, float)):
                    numeric_cols.append(h)
                    break
                elif str(val).replace(",", "").replace(".", "").replace("%", "").replace("$", "").strip().replace("-", "", 1).isdigit():
                    numeric_cols.append(h)
                    break
        if not numeric_cols:
            # Fall back to including ID-like columns rather than showing nothing,
            # in case EVERY numeric column happened to look ID-like.
            for h in headers:
                for row in real_data[:5]:
                    val = row.get(h)
                    if isinstance(val, (int, float)):
                        numeric_cols.append(h)
                        break
        if not numeric_cols:
            return # No numeric columns to summarize
        is_general = True
    else:
        is_general = False
        numeric_cols = []
        
    ws = wb.create_sheet("Dashboard", 1)  # Put it right after summary
    ws.sheet_view.showGridLines = False
    
    thin = Side(border_style="thin", color=theme["border"])
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    
    ws.merge_cells("A1:F2")
    title_text = "Executive Dashboard" if is_financial else ("Performance Dashboard" if is_performance else "Data Summary Dashboard")
    c = ws.cell(1, 1, title_text)
    c.font = Font(size=20, bold=True, color=theme["title_color"], name="Calibri")
    c.alignment = Alignment(vertical="center", horizontal="center")
    
    name_key = next((h for h in headers if any(word in str(h).lower() for word in ["name", "title", "company", "employee", "student", "item", "product"])), headers[0])

    if is_financial:
        _render_financial_dashboard(ws, real_data, headers, headers_lower, theme, brd, name_key)
    elif is_performance:
        _render_performance_dashboard(ws, real_data, headers, headers_lower, theme, brd, name_key)
    else:
        _render_general_dashboard(ws, real_data, headers, numeric_cols, theme, brd, name_key)

def _render_financial_dashboard(ws, real_data, headers, headers_lower, theme, brd, name_key):
    rev_key = next((h for h in headers if "revenue" in str(h).lower() or "sales" in str(h).lower()), None)
    prof_key = next((h for h in headers if "profit" in str(h).lower() and "margin" not in str(h).lower()), None)

    # Safety net: "is_financial" is detected from a broader keyword set
    # (revenue/profit/sales/margin/cost) than what this function actually
    # needs (revenue/sales OR profit specifically). If neither resolves —
    # e.g. the only match was "cost" or "margin" alone — building this P&L
    # table would silently produce an all-zero, meaningless dashboard.
    # Fall back to the generic numeric dashboard instead.
    if not rev_key and not prof_key:
        _id_like_words  = {"id", "no", "ref", "code", "num", "number", "reference"}
        _id_like_substr = ("serial", "uuid", "zipcode", "phone")
        numeric_cols = []
        for h in headers:
            if h == name_key:
                continue
            h_lower = str(h).strip().lower()
            words = re.split(r"[^a-z0-9]+", h_lower)
            if any(w in _id_like_words for w in words if w) or any(s in h_lower for s in _id_like_substr):
                continue
            for row in real_data[:5]:
                val = row.get(h)
                if isinstance(val, (int, float)):
                    numeric_cols.append(h)
                    break
        if numeric_cols:
            _render_general_dashboard(ws, real_data, headers, numeric_cols, theme, brd, name_key)
            return
        # No usable numeric column at all — nothing meaningful to show.
        ws.cell(4, 1, "No numeric columns found to analyze.").font = Font(italic=True, color=theme["sub_color"])
        return
    
    total_rev = 0
    total_prof = 0
    performers = []
    losses = 0
    
    for row in real_data:
        name = str(row.get(name_key, "Unknown"))
        rev = _safe_float(row.get(rev_key, 0)) if rev_key else 0
        prof = _safe_float(row.get(prof_key, 0)) if prof_key else 0
        
        total_rev += rev
        total_prof += prof
        
        if prof < 0: losses += 1
        performers.append({"name": name, "rev": rev, "prof": prof})
        
    avg_margin = (total_prof / total_rev) if total_rev else 0
    
    performers.sort(key=lambda x: x["prof"] if prof_key else x["rev"], reverse=True)
    
    top_perf = performers[0]["name"] if performers else "N/A"
    bot_perf = performers[-1]["name"] if performers else "N/A"
    
    kpis = []
    if rev_key: kpis.append(("Total Revenue", f"${total_rev:,.2f}"))
    if prof_key: kpis.append(("Net Profit", f"${total_prof:,.2f}"))
    if rev_key and prof_key: kpis.append(("Avg Margin %", f"{avg_margin*100:.1f}%"))
    kpis.append(("Top Performer", top_perf))
    kpis.append(("Lowest Performer", bot_perf))
    
    r = _render_kpi_section(ws, kpis, 4, theme, brd)
    
    insights = []
    total_entities = len(performers)
    if losses > 0:
        insights.append(f"⚠️ {losses} of {total_entities} entities operating at a loss.")
    else:
        insights.append(f"✅ All {total_entities} entities are profitable.")
        
    if avg_margin >= 0.15:
        insights.append(f"📈 Healthy margin of {avg_margin*100:.1f}% overall.")
    elif avg_margin > 0:
        insights.append(f"📊 Moderate margin of {avg_margin*100:.1f}% overall.")
    else:
        insights.append(f"📉 Negative or zero margin overall.")
        
    r = _render_insights_section(ws, insights, "Business Insights", r, theme)
    
    table_headers = ["Entity Name", "Revenue", "Profit", "Margin %", "Performance Rating"]
    r = _render_table_header(ws, table_headers, r, theme, brd)
    
    red_font = Font(color="9C0006")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    
    for p in performers:
        margin = p["prof"] / p["rev"] if p["rev"] else 0
        if margin >= 0.15: rating = "🟢 Strong"
        elif margin > 0: rating = "🟡 Moderate"
        else: rating = "🔴 Loss"
            
        cells = [
            ws.cell(r, 1, p["name"]),
            ws.cell(r, 2, p["rev"]),
            ws.cell(r, 3, p["prof"]),
            ws.cell(r, 4, margin),
            ws.cell(r, 5, rating)
        ]
        
        cells[1].number_format = "#,##0.00"
        cells[2].number_format = "#,##0.00"
        cells[3].number_format = "0.00%"
        
        for c in cells:
            c.border = brd
            c.alignment = Alignment(horizontal="center" if c.column > 1 else "left")
            if p["prof"] < 0:
                c.font = red_font
                c.fill = red_fill
            else:
                if r % 2 == 0: c.fill = PatternFill("solid", fgColor=theme["alt_row"])
        r += 1

def _render_performance_dashboard(ws, real_data, headers, headers_lower, theme, brd, name_key):
    score_key = next((h for h in headers if "score" in str(h).lower() or "mark" in str(h).lower()), None)
    if not score_key:
        score_key = next((h for h in headers if "rating" in str(h).lower() or "gpa" in str(h).lower()), None)
    
    scores = []
    performers = []
    
    for row in real_data:
        name = str(row.get(name_key, "Unknown"))
        score = _safe_float(row.get(score_key, 0)) if score_key else 0
        
        scores.append(score)
        performers.append({"name": name, "score": score})
        
    avg_score = sum(scores)/len(scores) if scores else 0
    max_score = max(scores) if scores else 0
    min_score = min(scores) if scores else 0
    
    performers.sort(key=lambda x: x["score"], reverse=True)
    
    top_perf = performers[0]["name"] if performers else "N/A"
    bot_perf = performers[-1]["name"] if performers else "N/A"
    
    kpis = [
        ("Avg Score", f"{avg_score:.1f}"),
        ("Highest Score", f"{max_score:g}"),
        ("Lowest Score", f"{min_score:g}"),
        ("Top Performer", top_perf),
        ("Lowest Performer", bot_perf),
    ]
    
    r = _render_kpi_section(ws, kpis, 4, theme, brd)
    
    insights = []
    total = len(performers)
    above_avg = sum(1 for p in performers if p["score"] > avg_score)
    insights.append(f"📊 Class / Group Average is {avg_score:.1f}.")
    insights.append(f"⭐ {above_avg} out of {total} scored above average.")
    
    r = _render_insights_section(ws, insights, "Performance Insights", r, theme)
    
    table_headers = ["Entity Name", "Score", "Status"]
    r = _render_table_header(ws, table_headers, r, theme, brd)
    
    red_font = Font(color="9C0006")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    
    for p in performers:
        score = p["score"]
        if score >= avg_score * 1.1: status = "🟢 Excellent"
        elif score >= avg_score * 0.9: status = "🟡 Good"
        else: status = "🔴 Needs Improvement"
            
        cells = [
            ws.cell(r, 1, p["name"]),
            ws.cell(r, 2, score),
            ws.cell(r, 3, status)
        ]
        
        cells[1].number_format = "General"
        
        for c in cells:
            c.border = brd
            c.alignment = Alignment(horizontal="center" if c.column > 1 else "left")
            if "Needs Improvement" in status:
                c.font = red_font
                c.fill = red_fill
            else:
                if r % 2 == 0: c.fill = PatternFill("solid", fgColor=theme["alt_row"])
        r += 1

def _render_general_dashboard(ws, real_data, headers, numeric_cols, theme, brd, name_key):
    metric_key = numeric_cols[0]
    
    vals = []
    items = []
    
    for row in real_data:
        name = str(row.get(name_key, "Unknown"))
        val = _safe_float(row.get(metric_key, 0))
        
        vals.append(val)
        items.append({"name": name, "val": val})
        
    total_val = sum(vals)
    avg_val = total_val / len(vals) if vals else 0
    max_val = max(vals) if vals else 0
    
    items.sort(key=lambda x: x["val"], reverse=True)
    
    top_item = items[0]["name"] if items else "N/A"
    
    kpis = [
        ("Total Entities", f"{len(items)}"),
        (f"Total {metric_key}", f"{total_val:g}"),
        (f"Avg {metric_key}", f"{avg_val:.1f}"),
        (f"Highest {metric_key}", top_item),
    ]
    
    r = _render_kpi_section(ws, kpis, 4, theme, brd)
    
    insights = []
    insights.append(f"📊 Dataset contains {len(items)} records.")
    insights.append(f"📈 Total {metric_key} is {total_val:g} with an average of {avg_val:.1f}.")
    
    r = _render_insights_section(ws, insights, "Data Insights", r, theme)
    
    table_headers = ["Entity Name", metric_key]
    r = _render_table_header(ws, table_headers, r, theme, brd)
    
    for p in items:
        cells = [
            ws.cell(r, 1, p["name"]),
            ws.cell(r, 2, p["val"])
        ]
        
        cells[1].number_format = "General"
        
        for c in cells:
            c.border = brd
            c.alignment = Alignment(horizontal="center" if c.column > 1 else "left")
            if r % 2 == 0: c.fill = PatternFill("solid", fgColor=theme["alt_row"])
        r += 1

def _render_kpi_section(ws, kpis, start_row, theme, brd):
    r = start_row
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(r, 1, "Key Performance Indicators").font = Font(bold=True, size=14, color=theme["title_color"])
    r += 1
    
    for i, (label, val) in enumerate(kpis):
        col = (i % 3) * 2 + 1
        row_idx = r + (i // 3) * 3
        
        ws.merge_cells(start_row=row_idx, start_column=col, end_row=row_idx, end_column=col+1)
        ws.cell(row_idx, col, label).font = Font(bold=True, color=theme["sub_color"])
        ws.cell(row_idx, col).fill = PatternFill("solid", fgColor="F0F4F8")
        ws.cell(row_idx, col).alignment = Alignment(horizontal="center")
        ws.cell(row_idx, col).border = brd
        
        ws.merge_cells(start_row=row_idx+1, start_column=col, end_row=row_idx+1, end_column=col+1)
        ws.cell(row_idx+1, col, val).font = Font(bold=True, size=14, color=theme["accent"])
        ws.cell(row_idx+1, col).alignment = Alignment(horizontal="center")
        ws.cell(row_idx+1, col).border = brd
        
    r += ((len(kpis)-1)//3 + 1) * 3 + 2
    return r

def _render_insights_section(ws, insights, title, r, theme):
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(r, 1, title).font = Font(bold=True, size=14, color=theme["title_color"])
    r += 1
    
    for text in insights:
        ws.merge_cells(f"A{r}:F{r}")
        ws.cell(r, 1, text).font = Font(italic=True, size=12)
        r += 1
        
    r += 2
    return r

def _render_table_header(ws, headers, r, theme, brd):
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(r, 1, "Detailed Analysis").font = Font(bold=True, size=14, color=theme["title_color"])
    r += 1
    
    for i, th in enumerate(headers, 1):
        c = ws.cell(r, i, th)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=theme["header_bg"])
        c.alignment = Alignment(horizontal="center")
        c.border = brd
    r += 1
    return r