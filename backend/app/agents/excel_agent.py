from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone

from app.models.state import AgentState, AgentStatus, SchemaType
from app.services.workspace_service import workspace_service
from app.services.ws_manager import ws_manager
import logging

logger = logging.getLogger(__name__)

THEME_COLORS = {
    "header_fill": "1F4E79",
    "header_font": "FFFFFF",
    "alt_row_fill": "EBF3FB",
    "border_color": "BDD7EE",
    "total_fill": "D6E4F0",
}


def _apply_header_style(cell):
    cell.font = Font(bold=True, color=THEME_COLORS["header_font"], size=11)
    cell.fill = PatternFill("solid", fgColor=THEME_COLORS["header_fill"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color=THEME_COLORS["border_color"])
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_data_style(cell, row_idx: int):
    if row_idx % 2 == 0:
        cell.fill = PatternFill("solid", fgColor=THEME_COLORS["alt_row_fill"])
    cell.alignment = Alignment(vertical="center")
    thin = Side(border_style="thin", color=THEME_COLORS["border_color"])
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


async def run_excel_agent(state: AgentState) -> AgentState:
    if state.status == AgentStatus.ERROR:
        return state

    await ws_manager.send_log(state.session_id, "ExcelAgent", "Generating Excel workbook...")

    try:
        data = state.cleaned_data or state.extracted_data or []
        if not data:
            state.error = "No data to write to Excel"
            state.status = AgentStatus.ERROR
            return state

        wb = Workbook()
        ws = wb.active
        ws.title = _get_sheet_title(state.schema_type)

        columns = list(data[0].keys())

        # Header row
        ws.row_dimensions[1].height = 30
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            _apply_header_style(cell)

        # Data rows
        for row_idx, row in enumerate(data, 2):
            ws.row_dimensions[row_idx].height = 18
            for col_idx, col_name in enumerate(columns, 1):
                value = row.get(col_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                _apply_data_style(cell, row_idx)

        # Auto-fit column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_len = max(
                len(str(col_name)),
                *[len(str(row.get(col_name, "") or "")) for row in data[:50]]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto filter
        ws.auto_filter.ref = ws.dimensions

        # Summary sheet
        _add_summary_sheet(wb, state, data, columns)

        # Save
        excels_dir = workspace_service.get_excels()
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        safe_name = (state.file_name or "output").replace(" ", "_").split(".")[0]
        output_path = excels_dir / f"{safe_name}_{timestamp}.xlsx"
        wb.save(output_path)

        state.output_excel_path = str(output_path)
        state.status = AgentStatus.COMPLETE

        await ws_manager.send_log(
            state.session_id, "ExcelAgent",
            f"Excel generated: {output_path.name} ({len(data)} rows)"
        )
        await ws_manager.send_status(state.session_id, AgentStatus.COMPLETE.value)

        return state

    except Exception as e:
        logger.error(f"ExcelAgent error: {e}")
        state.error = str(e)
        state.status = AgentStatus.ERROR
        return state


def _get_sheet_title(schema_type) -> str:
    if schema_type is None:
        return "Data"
    return {
        SchemaType.INVOICE: "Invoices",
        SchemaType.SALES_REPORT: "Sales",
        SchemaType.INVENTORY: "Inventory",
        SchemaType.HR_RECORDS: "HR Records",
        SchemaType.STUDENT_DATA: "Students",
        SchemaType.FINANCIAL: "Financial",
        SchemaType.GENERIC: "Data",
    }.get(schema_type, "Data")


def _add_summary_sheet(wb: Workbook, state: AgentState, data: list, columns: list):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30

    summary_data = [
        ("File", state.file_name or "N/A"),
        ("Schema Type", state.schema_type.value if state.schema_type else "generic"),
        ("Total Rows", len(data)),
        ("Total Columns", len(columns)),
        ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]

    ws.cell(row=1, column=1, value="SheetAgent AI — Processing Summary").font = Font(bold=True, size=13)

    for row_idx, (key, value) in enumerate(summary_data, 3):
        ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)
