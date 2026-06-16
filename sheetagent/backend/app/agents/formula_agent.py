"""
Formula Agent — Phase 2
Converts natural language requests into Excel formulas using Gemini,
then applies them to the generated workbook.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path
import re

from app.models.state import AgentState, AgentStatus
from app.services.gemini_service import gemini_service
from app.services.ws_manager import ws_manager
import logging

logger = logging.getLogger(__name__)

FORMULA_EXAMPLES = """
Natural language → Excel formula examples:
- "sum of sales column" → =SUM(B2:B100)
- "average price" → =AVERAGE(C2:C100)  
- "count orders" → =COUNT(A2:A100)
- "total if region is North" → =SUMIF(D2:D100,"North",B2:B100)
- "max revenue" → =MAX(E2:E100)
- "min cost" → =MIN(F2:F100)
- "profit margin" → =(B2-C2)/B2
- "lookup product name by ID" → =VLOOKUP(A2,Sheet2!A:B,2,FALSE)
- "unique count of categories" → =SUMPRODUCT(1/COUNTIF(D2:D100,D2:D100))
- "running total" → =SUM($B$2:B2)
- "top 10 filter" → =LARGE(B2:B100,1)
- "year from date" → =YEAR(A2)
- "if profit > 1000 show High else Low" → =IF(E2>1000,"High","Low")
"""


async def run_formula_agent(state: AgentState, formula_requests: list[str]) -> AgentState:
    """
    formula_requests: list of natural language strings like
    ["sum of revenue", "calculate profit margin", "show top 5 products"]
    """
    if not state.output_excel_path or not Path(state.output_excel_path).exists():
        await ws_manager.send_log(state.session_id, "FormulaAgent", "No Excel file found to apply formulas to", level="error")
        return state

    await ws_manager.send_log(state.session_id, "FormulaAgent", f"Processing {len(formula_requests)} formula request(s)...")

    try:
        wb = load_workbook(state.output_excel_path)
        ws_sheet = wb.active
        columns = [ws_sheet.cell(1, c).value for c in range(1, ws_sheet.max_column + 1)]
        last_row = ws_sheet.max_row
        last_col_letter = get_column_letter(ws_sheet.max_column)

        applied = []

        for request in formula_requests:
            await ws_manager.send_log(state.session_id, "FormulaAgent", f"Translating: '{request}'")

            formula_data = await _translate_to_formula(
                request, columns, last_row, last_col_letter
            )

            if formula_data.get("formula"):
                # Add formula to a new row after data
                formula_row = last_row + 2 + len(applied)
                label_cell = ws_sheet.cell(row=formula_row, column=1, value=formula_data.get("label", request))
                label_cell.font = Font(bold=True, color="1F4E79")

                result_cell = ws_sheet.cell(row=formula_row, column=2, value=formula_data["formula"])
                result_cell.font = Font(bold=True)
                result_cell.fill = PatternFill("solid", fgColor="D6E4F0")

                applied.append({
                    "request": request,
                    "formula": formula_data["formula"],
                    "label": formula_data.get("label", request),
                    "row": formula_row,
                })

                await ws_manager.send_log(
                    state.session_id, "FormulaAgent",
                    f"Applied: {formula_data['formula']}"
                )

        # Add a Formulas sheet summarizing all applied formulas
        if applied:
            _add_formulas_sheet(wb, applied)

        wb.save(state.output_excel_path)
        await ws_manager.send_log(state.session_id, "FormulaAgent", f"{len(applied)} formula(s) applied successfully")
        return state

    except Exception as e:
        logger.error(f"FormulaAgent error: {e}")
        await ws_manager.send_log(state.session_id, "FormulaAgent", f"Formula error: {e}", level="error")
        return state


async def _translate_to_formula(
    request: str,
    columns: list,
    last_row: int,
    last_col: str
) -> dict:
    prompt = f"""
You are an Excel formula expert. Convert the natural language request into an Excel formula.

Dataset columns: {columns}
Data rows: 2 to {last_row}

{FORMULA_EXAMPLES}

Request: "{request}"

Return JSON:
{{
  "formula": "=FORMULA(...)",
  "label": "Short descriptive label",
  "explanation": "What this formula does"
}}

Rules:
- Use actual column letters based on the columns list (A=first column, B=second, etc.)
- Always start formula with =
- Use row range 2:{last_row} for data ranges
- If request is unclear, return the best matching formula
- Never return null for formula if there's any reasonable interpretation
"""
    try:
        result = await gemini_service.analyze_json(prompt)
        return result if isinstance(result, dict) else {}
    except Exception as e:
        logger.error(f"Formula translation error: {e}")
        return {}


def _add_formulas_sheet(wb, applied: list):
    if "Formulas" in wb.sheetnames:
        del wb["Formulas"]

    ws = wb.create_sheet("Formulas")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20

    headers = ["Description", "Formula", "Applied At Row"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    for i, item in enumerate(applied, 2):
        ws.cell(row=i, column=1, value=item["label"])
        ws.cell(row=i, column=2, value=item["formula"])
        ws.cell(row=i, column=3, value=item["row"])
