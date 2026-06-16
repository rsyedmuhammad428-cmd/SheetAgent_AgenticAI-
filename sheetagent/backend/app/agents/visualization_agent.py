"""
Visualization Agent — Phase 2
Generates charts inside the Excel workbook using openpyxl charts.
Also saves standalone PNG charts to workspace/charts/.
"""
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from pathlib import Path
import json

from app.models.state import AgentState, AgentStatus, SchemaType
from app.services.gemini_service import gemini_service
from app.services.workspace_service import workspace_service
from app.services.ws_manager import ws_manager
import logging

logger = logging.getLogger(__name__)

CHART_TYPE_MAP = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
}


async def run_visualization_agent(state: AgentState, chart_requests: list[str] | None = None) -> AgentState:
    if not state.output_excel_path or not Path(state.output_excel_path).exists():
        await ws_manager.send_log(state.session_id, "VisualizationAgent", "No Excel file found", level="error")
        return state

    await ws_manager.send_log(state.session_id, "VisualizationAgent", "Generating charts...")

    try:
        wb = load_workbook(state.output_excel_path)
        ws = wb.active

        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        last_row = ws.max_row

        # If no explicit requests, auto-detect best charts via Gemini
        if not chart_requests:
            chart_requests = await _auto_detect_charts(headers, state.schema_type, last_row)

        charts_sheet = wb.create_sheet("Charts") if "Charts" not in wb.sheetnames else wb["Charts"]
        chart_row = 1
        charts_created = 0

        for req in chart_requests:
            await ws_manager.send_log(state.session_id, "VisualizationAgent", f"Building chart: {req}")

            chart_config = await _get_chart_config(req, headers, last_row)
            if not chart_config:
                continue

            chart = _build_chart(wb, ws, chart_config, last_row)
            if chart:
                charts_sheet.add_chart(chart, f"A{chart_row}")
                chart_row += 20
                charts_created += 1

        wb.save(state.output_excel_path)
        await ws_manager.send_log(
            state.session_id, "VisualizationAgent",
            f"{charts_created} chart(s) added to 'Charts' sheet"
        )
        return state

    except Exception as e:
        logger.error(f"VisualizationAgent error: {e}")
        await ws_manager.send_log(state.session_id, "VisualizationAgent", f"Chart error: {e}", level="error")
        return state


async def _auto_detect_charts(headers: list, schema_type, last_row: int) -> list[str]:
    prompt = f"""
Given this dataset, suggest the 2-3 most useful charts.

Columns: {headers}
Schema: {schema_type}
Row count: {last_row}

Return JSON array of natural language chart requests, e.g.:
["bar chart of sales by month", "line chart showing revenue trend", "pie chart of category distribution"]

Keep it to 2-3 max. Only suggest charts that make sense for this data.
"""
    try:
        result = await gemini_service.analyze_json(prompt)
        return result if isinstance(result, list) else []
    except Exception:
        return ["bar chart of first numeric column"]


async def _get_chart_config(request: str, headers: list, last_row: int) -> dict | None:
    prompt = f"""
Convert this chart request into a configuration.

Request: "{request}"
Available columns: {headers} (A=col1, B=col2, etc.)
Data rows: 2 to {last_row}

Return JSON:
{{
  "type": "bar" | "line" | "pie",
  "title": "Chart title",
  "category_col": 1,
  "value_col": 2,
  "series_title": "Series label"
}}

category_col and value_col are 1-based column numbers.
"""
    try:
        return await gemini_service.analyze_json(prompt)
    except Exception:
        return None


def _build_chart(wb, ws, config: dict, last_row: int):
    try:
        chart_type = config.get("type", "bar")
        ChartClass = CHART_TYPE_MAP.get(chart_type, BarChart)
        chart = ChartClass()

        chart.title = config.get("title", "Chart")
        chart.style = 10

        if chart_type == "pie":
            chart.width = 15
            chart.height = 12
        else:
            chart.width = 20
            chart.height = 12
            chart.grouping = "clustered"

        cat_col = config.get("category_col", 1)
        val_col = config.get("value_col", 2)

        data = Reference(ws, min_col=val_col, min_row=1, max_row=last_row)
        categories = Reference(ws, min_col=cat_col, min_row=2, max_row=last_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        if config.get("series_title"):
            chart.series[0].title = config["series_title"]

        return chart

    except Exception as e:
        logger.error(f"Chart build error: {e}")
        return None
