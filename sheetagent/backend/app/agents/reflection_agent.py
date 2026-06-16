"""
Reflection Agent — Phase 2
After Excel generation, verifies output quality and suggests fixes.
Mimics the 'reflection' step in agentic AI pipelines.
"""
from openpyxl import load_workbook
from pathlib import Path

from app.models.state import AgentState, AgentStatus
from app.services.gemini_service import gemini_service
from app.services.ws_manager import ws_manager
import logging

logger = logging.getLogger(__name__)


class ReflectionReport:
    def __init__(self):
        self.passed: list[str] = []
        self.warnings: list[str] = []
        self.errors: list[str] = []
        self.score: float = 0.0
        self.recommendation: str = ""

    def to_dict(self) -> dict:
        return {
            "passed": self.passed,
            "warnings": self.warnings,
            "errors": self.errors,
            "score": self.score,
            "recommendation": self.recommendation,
        }


async def run_reflection_agent(state: AgentState) -> tuple[AgentState, ReflectionReport]:
    report = ReflectionReport()

    await ws_manager.send_log(state.session_id, "ReflectionAgent", "Verifying output quality...")

    if not state.output_excel_path or not Path(state.output_excel_path).exists():
        report.errors.append("Output Excel file not found")
        report.score = 0.0
        return state, report

    try:
        # --- Structural checks ---
        wb = load_workbook(state.output_excel_path)
        ws = wb.active

        # Check 1: Has data
        if ws.max_row <= 1:
            report.errors.append("Excel file has no data rows")
        else:
            report.passed.append(f"Data present: {ws.max_row - 1} rows")

        # Check 2: Has headers
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        empty_headers = [i for i, h in enumerate(headers) if not h]
        if empty_headers:
            report.warnings.append(f"Empty column headers at positions: {empty_headers}")
        else:
            report.passed.append("All column headers present")

        # Check 3: Freeze panes
        if ws.freeze_panes:
            report.passed.append("Freeze panes applied")
        else:
            report.warnings.append("No freeze panes — consider freezing header row")

        # Check 4: Auto filter
        if ws.auto_filter.ref:
            report.passed.append("Auto-filter enabled")
        else:
            report.warnings.append("No auto-filter on headers")

        # Check 5: Summary sheet
        if "Summary" in wb.sheetnames:
            report.passed.append("Summary sheet present")
        else:
            report.warnings.append("No Summary sheet found")

        # Check 6: Data completeness via Gemini
        sample_rows = []
        for row in ws.iter_rows(min_row=2, max_row=min(10, ws.max_row), values_only=True):
            sample_rows.append(list(row))

        gemini_check = await _gemini_quality_check(headers, sample_rows, state.schema_type)
        report.warnings.extend(gemini_check.get("warnings", []))
        report.passed.extend(gemini_check.get("passed", []))
        report.recommendation = gemini_check.get("recommendation", "Output looks good.")

        # Score calculation
        total = len(report.passed) + len(report.warnings) + len(report.errors)
        if total > 0:
            report.score = round(len(report.passed) / total, 2)

        level = "info" if report.score >= 0.7 else "warning"
        await ws_manager.send_log(
            state.session_id, "ReflectionAgent",
            f"Quality score: {report.score:.0%} — {len(report.passed)} passed, {len(report.warnings)} warnings, {len(report.errors)} errors",
            level=level
        )

        # Append reflection to summary sheet
        _write_reflection_to_excel(wb, report, state.output_excel_path)

        return state, report

    except Exception as e:
        logger.error(f"ReflectionAgent error: {e}")
        report.errors.append(str(e))
        return state, report


async def _gemini_quality_check(headers: list, sample_rows: list, schema_type) -> dict:
    prompt = f"""
You are a data quality expert reviewing an Excel output.

Headers: {headers}
Schema type: {schema_type}
Sample rows (first 10): {sample_rows}

Check for:
1. Missing or null values in important columns
2. Inconsistent data formats
3. Suspicious outliers in numeric columns
4. Column names that are unclear or need renaming

Return JSON:
{{
  "passed": ["list of things that look good"],
  "warnings": ["list of specific data quality concerns"],
  "recommendation": "One sentence overall recommendation"
}}
"""
    try:
        return await gemini_service.analyze_json(prompt)
    except Exception:
        return {"passed": [], "warnings": [], "recommendation": "Could not perform AI quality check."}


def _write_reflection_to_excel(wb, report: ReflectionReport, path: str):
    if "Quality Report" in wb.sheetnames:
        del wb["Quality Report"]

    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet("Quality Report")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 55

    ws.cell(row=1, column=1, value="SheetAgent AI — Quality Report").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=f"Score: {report.score:.0%}").font = Font(bold=True, size=12)
    ws.cell(row=3, column=1, value=f"Recommendation: {report.recommendation}").font = Font(italic=True)

    row = 5
    for item in report.passed:
        ws.cell(row=row, column=1, value="✓ PASS").font = Font(color="1F7A1F", bold=True)
        ws.cell(row=row, column=2, value=item)
        row += 1

    for item in report.warnings:
        ws.cell(row=row, column=1, value="⚠ WARN").font = Font(color="B8860B", bold=True)
        ws.cell(row=row, column=2, value=item)
        row += 1

    for item in report.errors:
        ws.cell(row=row, column=1, value="✗ ERROR").font = Font(color="CC0000", bold=True)
        ws.cell(row=row, column=2, value=item)
        row += 1

    wb.save(path)
